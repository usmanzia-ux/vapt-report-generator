"""Render a Report to a styled multi-sheet Excel workbook (Cover / Findings /
Summary). Styling is adapted from the author's original VAPT register script.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import Report

HDR_BG = "0D1B2A"
HDR_FG = "FFFFFF"
ALT_ROW = "F2F6FA"

SEV_BG = {
    "Critical": "B00020",
    "High": "E65100",
    "Medium": "F9A825",
    "Low": "2E7D32",
    "Informational": "546E7A",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _border() -> Border:
    s = Side(style="thin", color="BFC9D4")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=HDR_FG, size=10)
    c.fill = _fill(HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()
    return c


def _cell(ws, row, col, value, bg=None, bold=False, color="1A2535"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=9, color=color)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = _border()
    if bg:
        c.fill = _fill(bg)
    return c


def _build_cover(wb: Workbook, report: Report) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = report.title.upper()
    t.font = Font(bold=True, size=15, color=HDR_FG)
    t.fill = _fill(HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    counts = report.severity_counts()
    rows = [
        ("Client", report.client),
        ("Assessor", report.assessor),
        ("Assessment Date", report.assessment_date),
        ("Standard", report.standard),
        ("Scope", ", ".join(report.scope) or "—"),
        ("Overall Risk", report.risk_rating()),
        ("Total Findings", str(len(report.findings))),
        ("Critical / High", f"{counts['Critical']} / {counts['High']}"),
        ("Medium / Low / Info", f"{counts['Medium']} / {counts['Low']} / {counts['Informational']}"),
    ]
    for i, (label, val) in enumerate(rows, start=3):
        c1 = ws.cell(row=i, column=1, value=label)
        c1.font = Font(bold=True, size=10)
        c1.fill = _fill("E8EEFF")
        c1.border = _border()
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=i, column=2, value=val)
        c2.font = Font(size=10)
        c2.border = _border()
        c2.alignment = Alignment(vertical="center")
        if label == "Overall Risk":
            c2.fill = _fill(SEV_BG.get(val, "FFFFFF"))
            c2.font = Font(size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[i].height = 20

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80


def _build_findings(wb: Workbook, report: Report) -> None:
    ws = wb.create_sheet("Findings")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    headers = [
        "ID", "Severity", "CVSS", "Finding", "Affected Targets",
        "Description", "CWE / CVE", "Remediation", "Source",
    ]
    widths = [9, 13, 8, 32, 30, 50, 18, 50, 12]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    th = ws["A1"]
    th.value = "DETAILED FINDINGS"
    th.font = Font(bold=True, size=13, color=HDR_FG)
    th.fill = _fill(HDR_BG)
    th.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(headers, 1):
        _hdr(ws, 2, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    ws.row_dimensions[2].height = 28

    for ri, f in enumerate(report.findings, start=3):
        targets = "\n".join(t.label() for t in f.targets) or "—"
        ids = " ".join(filter(None, [f.cwe or ""] + f.cve)) or "—"
        score = f"{f.cvss_score:.1f}" if f.cvss_score is not None else "—"
        vals = [
            f.finding_id, f.severity.value, score, f.title, targets,
            f.description, ids, f.remediation, f.source,
        ]
        row_bg = ALT_ROW if ri % 2 else "FFFFFF"
        for ci, val in enumerate(vals, 1):
            if ci == 2:  # severity colour
                _cell(ws, ri, ci, val, bg=SEV_BG.get(f.severity.value), bold=True, color="FFFFFF")
            else:
                _cell(ws, ri, ci, val, bg=row_bg)
        ws.row_dimensions[ri].height = 78

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(report.findings) + 2}"


def _build_summary(wb: Workbook, report: Report) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    s = ws["A1"]
    s.value = "SEVERITY SUMMARY"
    s.font = Font(bold=True, size=13, color=HDR_FG)
    s.fill = _fill(HDR_BG)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(["Severity", "Count", "% of Total"], 1):
        _hdr(ws, 2, ci, h)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14

    counts = report.severity_counts()
    total = max(len(report.findings), 1)
    for ri, sev in enumerate(["Critical", "High", "Medium", "Low", "Informational"], start=3):
        cnt = counts[sev]
        pct = f"{cnt / total * 100:.0f}%"
        _cell(ws, ri, 1, sev, bg=SEV_BG[sev], bold=True, color="FFFFFF")
        _cell(ws, ri, 2, cnt)
        _cell(ws, ri, 3, pct)
        ws.row_dimensions[ri].height = 20

    tr = 8
    _cell(ws, tr, 1, "TOTAL", bg="D6E4BC", bold=True)
    _cell(ws, tr, 2, len(report.findings), bg="D6E4BC", bold=True)
    _cell(ws, tr, 3, "100%", bg="D6E4BC", bold=True)


def render(report: Report, output: str) -> str:
    wb = Workbook()
    _build_cover(wb, report)
    _build_findings(wb, report)
    _build_summary(wb, report)
    wb.save(output)
    return output
